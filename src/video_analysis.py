#!/usr/bin/env python3
import sys
import os
import cv2
import numpy as np
import torch
import pickle as pkl
import mysql.connector
import datetime
from ultralytics import YOLO
from facenet_pytorch import InceptionResnetV1
from Crypto.Cipher import AES
from openpyxl import Workbook, load_workbook
from config import *
# ---------------------- CONFIGURATION ----------------------
KEY_FILE = "aes_secret.key"
INTRUDER_DIR = "Intruder_Log"
INTRUDER_XLS = "Intruder_Log.xlsx"
AUTH_XLS = "Authorized_Log.xlsx"

# Database Credentials
DB_HOST = "localhost"
DB_USER = "root"
DB_PASS = "YOUR_DB_PASSWORD"  # CAUTION: Do not use real passwords
DB_NAME = "face_db"

# Models & Thresholds
YOLO_FACE_MODEL = "yolov8s-face.pt"  
THRESHOLD = 0.85                     # Distance threshold (lower = stricter)
COOLDOWN = 60                        # Seconds between logging same authorized person
FACENET_DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# Ensure directories exist
os.makedirs(INTRUDER_DIR, exist_ok=True)
recent_log = {}

# ------------------ HELPER FUNCTIONS ------------------
def enhance_low_light(frame):
    """
    Applies CLAHE and Gamma Correction for better visibility.
    """
    try:
        lab = cv2.cvtColor(frame, cv2.COLOR_BGR2LAB)
        l, a, b = cv2.split(lab)
        cl = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8)).apply(l)
        limg = cv2.merge((cl, a, b))
        enhanced = cv2.cvtColor(limg, cv2.COLOR_LAB2BGR)
        
        gamma = 1.5
        lut = np.array([((i / 255.0) ** (1.0 / gamma)) * 255 for i in range(256)]).astype("uint8")
        return cv2.LUT(enhanced, lut)
    except Exception:
        return frame

def ensure_xls(path, headers):
    """
    Creates Excel file with headers if it doesn't exist.
    """
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(path)

def log_intruder(face_crop):
    """
    Logs intruder image and updates Excel.
    """
    ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    img_path = os.path.join(INTRUDER_DIR, f"intruder_{ts}.jpg")
    
    cv2.imwrite(img_path, face_crop)
    
    try:
        ensure_xls(INTRUDER_XLS, ["Timestamp", "Image Path"])
        wb = load_workbook(INTRUDER_XLS)
        ws = wb.active
        ws.append([ts, img_path])
        wb.save(INTRUDER_XLS)
    except PermissionError:
        print("⚠️ Close Intruder Excel to save log")

def log_authorized(name):
    """
    Logs authorized user with cooldown.
    """
    now = datetime.datetime.now()
    if name in recent_log:
        delta = (now - recent_log[name]).total_seconds()
        if delta < COOLDOWN:
            return
            
    recent_log[name] = now
    ts = now.strftime("%Y-%m-%d %H:%M:%S")
    
    try:
        ensure_xls(AUTH_XLS, ["Timestamp", "Name"])
        wb = load_workbook(AUTH_XLS)
        ws = wb.active
        ws.append([ts, name])
        wb.save(AUTH_XLS)
    except PermissionError:
        print("⚠️ Close Authorized Excel to save log")

# ------------------ DB & SECURITY ------------------
def load_aes_key(path=KEY_FILE):
    if not os.path.exists(path):
        raise FileNotFoundError(f"AES key not found: {path}")
    with open(path, "rb") as f:
        return f.read()

def load_authorized_embeddings(aes_key):
    out = {}
    rows = []
    
    try:
        cnx = mysql.connector.connect(
            host=DB_HOST, user=DB_USER, password=DB_PASS, database=DB_NAME
        )
        cur = cnx.cursor()
        cur.execute("SELECT name, embedding_cipher, aes_nonce, aes_tag FROM authorized_faces")
        rows = cur.fetchall()
        cur.close()
        cnx.close()
    except Exception as e:
        print(f"[ERROR] DB connect/fetch failed: {e}")
        return out

    for row in rows:
        try:
            name, cipher_blob, nonce, tag = row
            # Convert MySQL memoryview -> bytes if needed
            cipher_blob = bytes(cipher_blob)
            nonce = bytes(nonce)
            tag = bytes(tag)

            cipher = AES.new(aes_key, AES.MODE_GCM, nonce=nonce)
            plain = cipher.decrypt_and_verify(cipher_blob, tag)
            emb = pkl.loads(plain)
            emb = np.array(emb, dtype=np.float32).flatten()
            out[name] = emb
        except Exception as e:
            print(f"[WARN] Failed decrypt for {row[0] if row else '?'}: {e}")
            
    return out

# ------------------ RECOGNITION ------------------
def get_embedding_from_face(face_bgr, facenet_model):
    face_rgb = cv2.resize(face_bgr, (160, 160))
    face_rgb = cv2.cvtColor(face_rgb, cv2.COLOR_BGR2RGB)
    
    tensor = torch.tensor(face_rgb / 255., dtype=torch.float32).permute(2, 0, 1).unsqueeze(0).to(FACENET_DEVICE)
    
    with torch.no_grad():
        emb = facenet_model(tensor).cpu().numpy().flatten().astype(np.float32)
    return emb

def is_authorized(face_embedding, auth_dict, threshold=THRESHOLD):
    best_name = None
    best_dist = float("inf")
    
    for name, emb in auth_dict.items():
        d = np.linalg.norm(face_embedding - emb)
        if d < best_dist:
            best_dist = d
            best_name = name
            
    if best_name is not None and best_dist < threshold:
        return best_name, best_dist
        
    return None, best_dist

# ------------------ VIDEO PROCESSING ------------------
def process_video(video_path):
    if not os.path.exists(video_path):
        print(f"❌ Video not found: {video_path}")
        return

    print("[INFO] Loading models...")
    face_detector = YOLO(YOLO_FACE_MODEL)
    facenet_model = InceptionResnetV1(pretrained="vggface2").eval().to(FACENET_DEVICE)

    aes_key = load_aes_key(KEY_FILE)
    auth_embeds = load_authorized_embeddings(aes_key)
    print(f"[INFO] Loaded {len(auth_embeds)} authorized embeddings.")

    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        print(f"❌ Cannot open video: {video_path}")
        return

    print("[INFO] Processing video... Press 'q' to stop.")
    
    while True:
        ret, frame = cap.read()
        if not ret:
            break

        frame_enh = enhance_low_light(frame)
        results = face_detector(frame_enh, verbose=False)

        for r in results:
            boxes = getattr(r, "boxes", None)
            if boxes is None:
                continue
                
            for b in boxes:
                try:
                    # Handle different Ultralytics versions
                    x1, y1, x2, y2 = map(int, b.xyxy[0].tolist())
                except Exception:
                    arr = np.array(b.xyxy).reshape(-1)
                    x1, y1, x2, y2 = map(int, arr[:4])

                # Clamp coordinates
                h, w = frame_enh.shape[:2]
                x1, y1 = max(0, x1), max(0, y1)
                x2, y2 = min(w - 1, x2), min(h - 1, y2)

                if x2 <= x1 or y2 <= y1:
                    continue

                crop = frame_enh[y1:y2, x1:x2]
                if crop.size == 0:
                    continue

                emb = get_embedding_from_face(crop, facenet_model)
                name, dist = is_authorized(emb, auth_embeds, threshold=THRESHOLD)

                if name:
                    log_authorized(name)
                    color = (0, 255, 0)
                    label = f"{name} ({dist:.2f})"
                else:
                    log_intruder(crop)
                    color = (0, 0, 255)
                    label = f"INTRUDER ({dist:.2f})"

                cv2.rectangle(frame_enh, (x1, y1), (x2, y2), color, 2)
                cv2.putText(frame_enh, label, (x1, max(12, y1 - 10)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

        cv2.imshow("Video Check", frame_enh)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()
    print("[INFO] Video processing complete.")

# ------------------ ENTRY POINT ------------------
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 video_check.py /path/to/video.mp4")
        sys.exit(1)
        
    video_path = sys.argv[1]
    process_video(video_path)