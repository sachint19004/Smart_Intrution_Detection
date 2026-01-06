import os
import cv2
import socket
import threading
import datetime
import numpy as np
import torch
import pickle as pkl
import mysql.connector
from ultralytics import YOLO
from facenet_pytorch import InceptionResnetV1
from Crypto.Cipher import AES
from openpyxl import Workbook, load_workbook
from config import *
# ---------------- CONFIGURATION ----------------
HOST = "0.0.0.0"  # Listen on all interfaces
PORT = 5000
KEY_FILE = "aes_secret.key"
COOLDOWN = 60  # Seconds between logging same identity

# Logging Paths
INTRUDER_DIR = "Intruder_Log"
INTRUDER_XLS = "Intruder_Log.xlsx"
AUTH_XLS = "Authorized_Log.xlsx"

# Database Configuration
DB_HOST = "localhost"
DB_USER = "root"
DB_PASS = "YOUR_DB_PASSWORD"  
DB_NAME = "face_db"

# Models & Thresholds
YOLO_FACE = "yolov8s-face.pt"  
FACENET_PRETRAIN = "vggface2"
COSINE_THRESHOLD = 0.75
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"

# Ensure directories exist
os.makedirs(INTRUDER_DIR, exist_ok=True)

# ---------------- INITIALIZATION ----------------
print(f"[INFO] Using device: {DEVICE}")
face_model = YOLO(YOLO_FACE)
facenet = InceptionResnetV1(pretrained=FACENET_PRETRAIN).eval().to(DEVICE)

# Load AES Key
if not os.path.exists(KEY_FILE):
    raise FileNotFoundError(f"[ERROR] AES key file not found: {KEY_FILE}")

with open(KEY_FILE, "rb") as f:
    AES_KEY = f.read()

# ---------------- LOGGING HELPERS ----------------
recent_log = {}  # Cache: name -> last datetime logged

def write_xls_row(path, row, header):
    """
    Appends a row to an Excel file; creates it if missing.
    """
    try:
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.append(header)
            wb.save(path)
        
        wb = load_workbook(path)
        ws = wb.active
        ws.append(row)
        wb.save(path)
    except PermissionError:
        print(f"[WARN] Cannot write to {path}. Close file if open in Excel.")

def log_intruder(face_crop):
    """
    Saves the intruder's face image and logs to Excel.
    """
    ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    img_path = os.path.join(INTRUDER_DIR, f"intruder_{ts}.jpg")
    
    try:
        cv2.imwrite(img_path, face_crop)
    except Exception as e:
        print(f"[WARN] Failed saving intruder image: {e}")
        return

    write_xls_row(INTRUDER_XLS, [ts, img_path], ["timestamp", "image_path"])
    print(f"[INFO] Intruder logged: {img_path}")

def log_authorized(name):
    """
    Logs authorized access with a cooldown to prevent spamming.
    """
    now = datetime.datetime.now()
    
    if name in recent_log:
        delta = (now - recent_log[name]).total_seconds()
        if delta < COOLDOWN:
            return

    recent_log[name] = now
    ts = now.strftime("%Y-%m-%d %H:%M:%S")
    write_xls_row(AUTH_XLS, [ts, name], ["timestamp", "name"])
    print(f"[INFO] Authorized: {name} at {ts}")

# ---------------- DATABASE & EMBEDDINGS ----------------
def load_authorized_embeddings():
    """
    Fetch, decrypt, and normalize embeddings from the database.
    Returns: dict {name: [numpy_embedding_array, ...]}
    """
    auth = {}
    rows = []
    
    try:
        conn = mysql.connector.connect(
            host=DB_HOST, user=DB_USER, password=DB_PASS, database=DB_NAME
        )
        cur = conn.cursor()
        cur.execute("SELECT name, embedding_cipher, aes_nonce, aes_tag FROM authorized_faces")
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[ERROR] DB connection failed: {e}")
        return auth

    for row in rows:
        try:
            name, cipher_blob, nonce, tag = row
            cipher = AES.new(AES_KEY, AES.MODE_GCM, nonce=nonce)
            decrypted = cipher.decrypt_and_verify(cipher_blob, tag)
            
            emb = pkl.loads(decrypted)
            emb = np.array(emb, dtype=np.float32)
            
            # Normalize embedding
            emb /= (np.linalg.norm(emb) + 1e-12)
            
            auth.setdefault(name, []).append(emb)
        except Exception as e:
            print(f"[WARN] Decryption failed for user {row[0]}: {e}")
            continue

    print(f"[INFO] Loaded embeddings for {len(auth)} identities")
    return auth

def get_embedding_from_face(face_bgr):
    """
    Converts a BGR face crop to a normalized embedding vector.
    """
    try:
        face_rgb = cv2.resize(face_bgr, (160, 160))
        face_rgb = cv2.cvtColor(face_rgb, cv2.COLOR_BGR2RGB)
        
        tensor = torch.tensor(face_rgb / 255.0, dtype=torch.float32).permute(2, 0, 1).unsqueeze(0).to(DEVICE)
        
        with torch.no_grad():
            emb = facenet(tensor).cpu().numpy().flatten().astype(np.float32)
        
        emb /= (np.linalg.norm(emb) + 1e-12)
        return emb
    except Exception as e:
        print(f"[WARN] Embedding generation failed: {e}")
        return None

def cosine_similarity(a, b):
    return float(np.dot(a, b) / ((np.linalg.norm(a) * np.linalg.norm(b)) + 1e-12))

def recognize_face(emb, auth_embeds, threshold=COSINE_THRESHOLD):
    """
    Compares the input embedding against the authorized database.
    Returns: (name, score) or (None, best_score)
    """
    best_name = None
    best_score = -1.0

    for name, emb_list in auth_embeds.items():
        for ref in emb_list:
            s = cosine_similarity(emb, ref)
            if s > best_score:
                best_score = s
                best_name = name
    
    if best_score >= threshold:
        return best_name, best_score
    
    return None, best_score

# ---------------- IMAGE PROCESSING ----------------
def enhance_low_light(frame):
    """
    Applies CLAHE and Gamma Correction for better visibility in dark conditions.
    """
    try:
        # CLAHE on L channel
        lab = cv2.cvtColor(frame, cv2.COLOR_BGR2LAB)
        l, a, b = cv2.split(lab)
        cl = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8)).apply(l)
        limg = cv2.merge((cl, a, b))
        enhanced = cv2.cvtColor(limg, cv2.COLOR_LAB2BGR)

        # Gamma Correction (Gamma = 1.5)
        gamma = 1.5
        lut = np.array([((i / 255.0) ** (1.0 / gamma)) * 255 for i in range(256)]).astype("uint8")
        return cv2.LUT(enhanced, lut)
    except Exception:
        return frame  # Return original if enhancement fails

# ---------------- NETWORK HANDLER ----------------
def handle_connection(conn, addr, auth_embeds):
    """
    Threaded function to handle incoming video stream from a client.
    """
    print(f"[INFO] Connection from {addr}")
    data = b""
    payload_size = 4  # 4 bytes for integer size

    try:
        while True:
            # 1. Read message header (Frame Size)
            while len(data) < payload_size:
                packet = conn.recv(4096)
                if not packet:
                    raise ConnectionError("Client closed connection")
                data += packet
            
            msg_size = int.from_bytes(data[:payload_size], "big")
            data = data[payload_size:]

            # 2. Read frame payload
            while len(data) < msg_size:
                packet = conn.recv(4096)
                if not packet:
                    raise ConnectionError("Client closed connection")
                data += packet
            
            frame_data = data[:msg_size]
            data = data[msg_size:]

            # 3. Decode Frame
            try:
                buffer = pkl.loads(frame_data)
                frame = cv2.imdecode(buffer, cv2.IMREAD_COLOR)
            except Exception as e:
                print(f"[WARN] Frame decode failed: {e}")
                continue
            
            if frame is None:
                continue

            # 4. Processing
            frame_enh = enhance_low_light(frame)
            results = face_model(frame_enh, verbose=False)

            for res in results:
                for box in res.boxes.xyxy.cpu().numpy():
                    x1, y1, x2, y2 = map(int, box)
                    
                    # Clamp coordinates
                    h, w = frame_enh.shape[:2]
                    x1, y1 = max(0, x1), max(0, y1)
                    x2, y2 = min(w - 1, x2), min(h - 1, y2)

                    if x2 <= x1 or y2 <= y1:
                        continue
                    
                    face_crop = frame_enh[y1:y2, x1:x2]
                    if face_crop.size == 0:
                        continue

                    # Generate Embedding
                    emb = get_embedding_from_face(face_crop)
                    if emb is None:
                        continue

                    # Recognition
                    match_name, score = recognize_face(emb, auth_embeds)
                    
                    if match_name:
                        log_authorized(match_name)
                        label = f"{match_name} {score:.2f}"
                        color = (0, 255, 0) # Green
                    else:
                        log_intruder(face_crop)
                        label = f"INTRUDER {score:.2f}"
                        color = (0, 0, 255) # Red

                    # Draw UI
                    cv2.rectangle(frame_enh, (x1, y1), (x2, y2), color, 2)
                    cv2.putText(frame_enh, label, (x1, max(20, y1 - 8)),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)

            # Display Feed
            cv2.imshow(f"Receiver - {addr}", frame_enh)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                print("[INFO] Quit requested by user")
                return

    except (ConnectionError, KeyboardInterrupt) as e:
        print(f"[INFO] Connection ended ({addr}): {e}")
    finally:
        try:
            conn.close()
        except:
            pass
        cv2.destroyAllWindows()

# ---------------- MAIN SERVER ----------------
def start_receiver():
    # Load embeddings once at start
    auth_embeds = load_authorized_embeddings()
    
    server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    
    try:
        server.bind((HOST, PORT))
        server.listen(5)
        print(f"[INFO] Listening on {HOST}:{PORT} ... (Press Ctrl+C to stop)")
        
        while True:
            conn, addr = server.accept()
            # Spawn a thread per client
            t = threading.Thread(target=handle_connection, args=(conn, addr, auth_embeds), daemon=True)
            t.start()
            
    except KeyboardInterrupt:
        print("\n[INFO] Server stopped by user")
    finally:
        server.close()
        cv2.destroyAllWindows()

if __name__ == "__main__":
    start_receiver()